import tkinter as tk
import datetime as dt
import locale
import os
import openpyxl as xl
import csv
import datetime as dt
import calendar

from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkcalendar import DateEntry
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from mother import path

# Set locale menjadi bahasa Indonesia
locale.setlocale(locale.LC_TIME, 'id_ID')

class VisualDataManagement():
    def __init__(self, inface, raw_attd) -> None:
        self.inface = inface
        self.raw_attd = raw_attd
        self.add_button = tk.Button(self.inface.frame2, text='Tambah', command=self.new_window_list_files)
        self.boxs = []
        self.wb = xl.Workbook()
        self.ws = self.wb.active
        self.refresh_cells()
        self.color = {
            "morning" : "92D050",
            "afternoon" : "00B0F0",
            "r_morning" : "FF0000",
            "r_afternoon" : "F4B084"
        }

    def generate_excel(self):
        # jika 
        if len(self.raw_attd.primary_data) == 0:
            messagebox.showwarning('Daftar Sekolah Kosong', "Silakan nama sekolah disi terlebih dahulu")
            self.raw_attd.open_file('text')
            return
        
        now = dt.datetime.now()
        folder_path = filedialog.asksaveasfilename(defaultextension='xlsx', initialfile=str(now.strftime('%B')))
        if not folder_path:
            return

        file_names = os.listdir(path['fix'])
        cek_files = []
        # mengisi array cek_files
        for cell in self.boxs:
            for file in file_names:
                if cell["file_name"] == file:
                    cek_files.append(True)
        
        # mengecek apakah semua cell sudah memiliki file absen
        if len(cek_files) == len(self.boxs):
            print("semua lengkap")
        else:
            messagebox.showwarning('Tidak Lengkap', "Ada cell yang tidak memiliki file absen")
            return
        
        # membuat worbook excel
        self.wb = xl.Workbook()
        self.ws = self.wb.active
        ws = self.ws

        ws['A1'] = "NO"
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = self.border()
        
        ws["B1"] = "NAMA SEKOLAH"
        ws['B1'].border = self.border()
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

        # mengisi cell utama
        for index, school in enumerate(self.raw_attd.primary_data):
            cell_A = ws['A'+str(index+2)]
            cell_B = ws['B'+str(index+2)]
            cell_A.value = index+1
            cell_A.border = self.border()
            cell_B.value = str(school)
            cell_B.border = self.border()

        ws.column_dimensions['B'].width = self.get_len_cell_in_column(ws, 'B')+2
        ws.column_dimensions['A'].width = self.get_len_cell_in_column(ws, 'A')+2
        ws.row_dimensions[1].height = 50

        # mengisi warna cell
        for index, file_raw in enumerate(self.boxs):
            file = file_raw['file_name']
            cell_name = chr(ord('C')+index*2)
            ncell_name = chr(ord(cell_name)+1)
            cellbox = self.get_cellbox_with_file_name(file)
            
            self.ws[cell_name+'1'].alignment = Alignment(horizontal='center', vertical='center')
            self.ws[ncell_name+'1'].alignment = Alignment(horizontal='center', vertical='center')
            self.ws[cell_name+'1'].border = self.border()
            self.ws[ncell_name+'1'].border = self.border()

            ws.column_dimensions[cell_name].width = 15
            ws.column_dimensions[ncell_name].width = 15
            
            self.ws[cell_name+'1'] = cellbox['date'].strftime("%A %d %b \n%Y")+" Pagi"
            self.ws[ncell_name+'1'] = cellbox['date'].strftime("%A %d %b \n%Y")+" Sore"

            self.fill_2column_color(cell_name, self.color["r_morning"], self.color['r_afternoon'])

            with open(path['fix']+file, 'r') as csv_reader:
                csv_reader = csv.reader(csv_reader, delimiter=',', quotechar='"')
                for row in csv_reader:
                    c_pos = self.get_position_school(row[2])
                    if c_pos:
                        c_pos += 1
                        if row[4] == 'pagi':
                            cell = ws[cell_name+str(c_pos)]
                            cell.fill = PatternFill(start_color=self.color['morning'], end_color=self.color['morning'], fill_type="solid")
                        else:
                            cell = ws[chr(ord(cell_name)+1)+str(c_pos)]
                            cell.fill = PatternFill(start_color=self.color['afternoon'], end_color=self.color['afternoon'], fill_type="solid")
            
        self.wb.save(folder_path)
        messagebox.showinfo("Berhasil!!!", "Berhasil membuat absen excel!!!")
        
    def get_position_school(self, attendance_school_name):
        for index, school in enumerate(self.raw_attd.primary_data):
            if self.raw_attd.normalize_school_name(attendance_school_name) == self.raw_attd.normalize_school_name(school):
                return index+1
    
    def get_len_cell_in_column(self, ws, column):
        cell_len = []
        for cell in ws[column]:
            cell_len.append(len(str(cell.value)))
        return max(cell_len)

    def create_cell(self, year, month, day, filename=''):
        date = dt.datetime(year, month, day)
        return {
            "cell" : tk.Label(self.inface.frame2, text=filename+'\n'+date.strftime("%A\n%d %b %Y"), background='#897d76'),
            "file_name" : filename,
            "date" : date
        }
    
    def refresh_cells(self):
        rowcolumn = self.generate_rows_columns(len(self.boxs)+1, 4)

        for i, value in enumerate(rowcolumn):
            if i == len(self.boxs):
                self.add_button.grid(row=value[0], column=value[1], sticky='wens', padx=5, pady=5)
            else:
                date = self.boxs[i]['date']
                file_name = self.boxs[i]['file_name']
                self.boxs[i]["cell"].grid(row=value[0], column=value[1], sticky='wens', padx=5, pady=5)
                self.boxs[i]["cell"].bind('<Button-3>', self.cell_bind)
                self.boxs[i]['cell'].config(text=file_name+'\n'+date.strftime("%A\n%d %b %Y"))

    def cell_bind(self, event):
        self.cell_menu(event)
    
    def cell_menu(self, event):
        menu = tk.Menu(self.inface.window, tearoff=0)
        menu.add_command(label="Change", command=lambda: self.new_window(event))
        menu.add_command(label="Delete", command=lambda:self.delete_cell(event))
        menu.add_command(label="Delete All", command=lambda:self.delete_cells(event))
        menu.add_command(label="Generate", command=self.generate_excel)
        menu.post(event.x_root, event.y_root)
        
    def delete_cell(self, event):
        result = messagebox.askyesno("Konfirmasi", "Apakah Anda yakin?")
        if result:
            widget = event.widget
            cell = self.get_cell(event)
            if cell:
                print(cell, " dihapus")
                self.boxs.remove(cell)
                cell['cell'].destroy()
            self.refresh_cells()
    
    def delete_cells(self, event):
        result = messagebox.askyesno("Konfirmasi", "Apakah Anda yakin menghapus semuanya?")
        if result:
            for cell in self.boxs:
                cell['cell'].destroy()
            self.boxs.clear()
            self.refresh_cells()

    def new_window_list_files(self):
        new_window = tk.Toplevel(self.inface.window)
        new_window.bind("<Return>", lambda event: self.add_cell(combo_box.get(), new_window))
        new_window.geometry(f'+{self.inface.window.winfo_pointerx()}+{self.inface.window.winfo_pointery()}')
        new_window.resizable(0, 0)

        files = []
        for file in os.listdir(path['fix']):
            files.append(str(file))

        combo_box = ttk.Combobox(new_window, value=files)
        button = ttk.Button(new_window, text="Add_All", command=lambda: self.add_cells(new_window))
        button.pack()
        combo_box.pack()
    
    def add_cell(self, combo_box_value, new_window):
        if len(combo_box_value.split(' ')) > 1:
            month = combo_box_value.split(' ')[0].replace(' ', '')
            date = combo_box_value.split(' ')[1].split('.')[0]
            datetime = dt.datetime.now()
        for idx, m in enumerate(calendar.month_name):
            if date and month and m == month:
                datetime = dt.datetime(dt.datetime.now().year, idx, int(date))

        self.boxs.append(self.create_cell(datetime.year, datetime.month, datetime.day, combo_box_value))
        
        self.refresh_cells()
        new_window.destroy()
    
    def add_cells(self, new_window):
        files = []
        for file in os.listdir(path['fix']):
            files.append(str(file))
            
        for file in files:
            month = file.split(' ')[0].replace(' ', '')
            date = file.split(' ')[1].split('.')[0]
            datetime = dt.datetime.now()
            for idx, m in enumerate(calendar.month_name):
                if date and month and m == month:
                    datetime = dt.datetime(dt.datetime.now().year, idx, int(date))

            self.boxs.append(self.create_cell(datetime.year, datetime.month, datetime.day, file))
        self.refresh_cells()
        new_window.destroy()

    def new_window(self, event):
        cell = self.get_cell(event)
        new_window = tk.Toplevel(self.inface.window)
        new_window.geometry(f'+{event.x_root}+{event.y_root}')

        files = []
        for file in os.listdir(path['fix']):
            files.append(str(file))
        
        combo_box = ttk.Combobox(new_window, value=files)
        date_entry = DateEntry(new_window)
        button_save = tk.Button(new_window, text="Save", command=lambda: self.save_change(event, combo_box, date_entry, new_window))
        
        # memasukan nilai di inputan
        if cell:
            if cell['file_name']:
                combo_box.set(cell['file_name'])
                date_entry.set_date(cell['date'])

        combo_box.pack()
        date_entry.pack()
        button_save.pack()
    
    def get_cell(self, event):
        widget = event.widget
        for i in self.boxs:
            if i['cell'].info == widget.info:
                return i
    
    def get_cellbox_with_file_name(self, file_name):
        for i in self.boxs:
            if i['file_name'] == file_name:
                return i
        
    def save_change(self, event, combo_box, date_entry, new_window):
        cell = self.get_cell(event)
        if cell:
            cell['file_name'] = combo_box.get()
            cell['date'] = date_entry.get_date()
        self.refresh_cells()
        new_window.destroy()

    def generate_rows_columns(self, length, limit):
        result = []
        for i in range(length):
            row = i // limit
            column = i % limit
            result.append([row, column])
        return result

    def calculate_factors(self, product):
        factors = []
        for i in range(1, int(product ** 0.5) + 1):
            if product % i == 0:
                factors.append((i, product // i))
        
        # Mengembalikan pasangan faktor dengan nilai terkecil
        smallest_factors = min(factors, key=lambda x: sum(x))
        return smallest_factors

    def border(self):
        return Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
    
    def fill_2column_color(self, column_name, c_column1, c_column2):
        # membuat cell menjadi merah
        for index in range(len(self.raw_attd.primary_data)):
            # cell pagi
            cell = self.ws[column_name+str(index+2)]
            cell.fill = PatternFill(start_color=c_column1, end_color=c_column1, fill_type="solid")
            cell.border = self.border()
            # cell sore
            cell = self.ws[chr(ord(column_name)+1)+str(index+2)]
            cell.fill = PatternFill(start_color=c_column2, end_color=c_column2, fill_type="solid")
            cell.border = self.border()
        
